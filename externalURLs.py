import requests
from openpyxl import load_workbook, Workbook
import os 

INPUT_EXCEL_FILE = "dashboard_guids.xlsx"  
OUTPUT_EXCEL_FILE = "dashboard_urls.xlsx"  

endpoint = "https://api.newrelic.com/graphql"
api_key = os.environ.get('API_KEY')

def read_dashboard_data(file_path):
    try:
        print(f"Attempting to read file: {file_path}")
        wb = load_workbook(filename=file_path)
        ws = wb.active

        dashboards = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            guid, name = row[:2]  
            if guid and name:
                dashboards.append((guid, name))

        print(f"{len(dashboards)} Dashboards found.") if dashboards else print("No Dashboards found in the file!")
        return dashboards
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found!")
        return []
    except Exception as e:
        print(f"Error reading data from Excel: {e}")
        return []

def fetch_dashboard_url(guid):
    query = f"""
    mutation {{
      dashboardCreateSnapshotUrl(guid: "{guid}")
    }}
    """
    try:
        response = requests.post(endpoint, 
                                 json={"query": query}, 
                                 headers={"API-Key": api_key,
                                          "Content-Type": "application/json"})
        response.raise_for_status()

        data = response.json()
        url = data.get("data", {}).get("dashboardCreateSnapshotUrl", "")

        if not url:
            print(f"No URL found for GUID {guid}")
        return url
    except requests.exceptions.RequestException as e:
        print(f"API Request Failed for GUID {guid}: {e}")
        return "ERROR"

def save_dashboard_urls_to_excel(dashboard_data, urls):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard URLs"

    ws.append(["Dashboard Name", "GUID", "Dashboard URL"])

    for (guid, name), url in zip(dashboard_data, urls):
        ws.append([name, guid, url])

    wb.save(OUTPUT_EXCEL_FILE)
    print(f"Dashboard URLs saved successfully in '{OUTPUT_EXCEL_FILE}'")

def main():
    print("============Reading Dashboards from Excel============")
    dashboard_data = read_dashboard_data(INPUT_EXCEL_FILE)

    if not dashboard_data:
        print("============No Dashboards found. Please check the file structure and data============")
        return

    print("============Fetching dashboard URLs============")
    urls = [fetch_dashboard_url(guid) for guid, _ in dashboard_data]

    print("============Saving URLs to Excel============")
    save_dashboard_urls_to_excel(dashboard_data, urls)

if __name__ == "__main__":
    main()
