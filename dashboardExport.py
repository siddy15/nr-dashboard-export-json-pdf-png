import requests
import json
from openpyxl import load_workbook
import re, os

# Configuration
endpoint = "https://api.newrelic.com/graphql" 

# Set the API_key as environment variable
api_key = os.environ.get('API_KEY')

# Path to the Excel file
input_excel = "dashboard_guids.xlsx"  

query_file = "dashboardExportQuery.graphql"

# Loading the GraphQL
def load_query_template(file_path):
    try:
        with open(file_path, "r") as file:
            return file.read()
    except FileNotFoundError:
        print(f"Error: The query file '{file_path}' does not exist.")
        exit(1)

# Iterating over guids from excel sheet
def load_guids_from_excel(file_path):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        return [row[0].value for row in sheet.iter_rows(min_row=2, max_col=1) if row[0].value]
    except FileNotFoundError:
        print(f"Error: The Excel file '{file_path}' does not exist.")
        exit(1)
    except Exception as e:
        print(f"Error: Unable to read the Excel file. Details: {e}")
        exit(1)

# Standardizing the file names 
def sanitize_filename(name):
    return re.sub(r'[<>:"/\\|?*\n]', '_', name)


# Function to fetch dashboard data for a given GUID
def fetch_dashboard_data(guid, query_template):
    query = query_template % guid
    try:
        response = requests.post(
            endpoint,
            json={"query": query},
            headers={"API-Key": api_key, "Content-Type": "application/json"},
        )
        if response.status_code == 200:
            return response.json()
        else:
            print(f"Failed to fetch data for GUID {guid}. HTTP Status Code: {response.status_code}")
            return None
    except requests.RequestException as e:
        print(f"Error: Failed to fetch data for GUID {guid}. Details: {e}")
        return None


# Function to save data to a JSON file
def save_data_to_file(data, dashboard_name):
    sanitized_name = sanitize_filename(dashboard_name)
    file_name = f"{sanitized_name}.json"
    try:
        with open(file_name, "w") as file:
            json.dump(data, file, indent=2)
        print(f"Data saved successfully in {file_name}")
    except FileNotFoundError:
        print(f"Error: Could not create file for dashboard '{dashboard_name}'. Skipping...")


# Main function to orchestrate the workflow
def main():
    # Load query template
    print("========== Loading query template ==========")
    query_template = load_query_template(query_file)

    # Load GUIDs from Excel
    print("========== Loading GUIDs from Excel ==========")
    guids = load_guids_from_excel(input_excel)
    print(f"Found {len(guids)} GUIDs to process.")

    # Iterate over GUIDs and fetch data
    for guid in guids:
        print(f"Processing GUID: {guid}")
        data = fetch_dashboard_data(guid, query_template)
        if data:
            try:
                # Extract dashboard name
                dashboard_name = data["data"]["actor"]["entity"]["name"]
                save_data_to_file(data, dashboard_name)
            except KeyError:
                print(f"Error: Unable to find the dashboard name for GUID {guid}.")

# Entry point
if __name__ == "__main__":
    main()
